import io
from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .models import Cuota, Pago, AlertaSeguridad

def generar_reporte_finanzas_excel():
    """Genera un archivo Excel profesional con el reporte financiero"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Financiero"
    
    # --- Estilos ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    
    border_style = Side(border_style="thin", color="000000")
    border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
    
    currency_format = 'Bs #,##0.00'

    # --- Encabezado ---
    headers = ["ID", "Residente", "Unidad", "Mes", "Monto Cuota", "Estado", "Total Pagado", "Saldo Pendiente"]
    ws.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # --- Datos ---
    cuotas = Cuota.objects.select_related('residente__user', 'residente__unidad_habitacional').all().order_by('-id')
    
    for c in cuotas:
        total_pagado = sum(p.monto_pagado for p in c.pagos.all())
        saldo = c.monto - total_pagado
        
        row = [
            c.id,
            c.residente.user.get_full_name(),
            str(c.residente.unidad_habitacional),
            c.mes,
            c.monto,
            c.estado.upper(),
            total_pagado,
            saldo
        ]
        ws.append(row)
        
        # Formato de celda por fila
        current_row = ws.max_row
        ws.cell(row=current_row, column=5).number_format = currency_format # Monto
        ws.cell(row=current_row, column=7).number_format = currency_format # Pagado
        ws.cell(row=current_row, column=8).number_format = currency_format # Saldo

        # Colorear estado
        state_cell = ws.cell(row=current_row, column=6)
        state_cell.alignment = center_align
        if c.estado == 'pagada':
            state_cell.font = Font(color="008000", bold=True)
        elif c.estado == 'vencida':
            state_cell.font = Font(color="FF0000", bold=True)

    # --- Ajustar ancho de columnas ---
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Output buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generar_reporte_seguridad_pdf():
    """Genera un PDF profesional con tablas usando ReportLab Platypus"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    title_style.alignment = 1 # Center
    
    # --- Título ---
    elements.append(Paragraph("Reporte de Seguridad - Smart Condominium", title_style))
    elements.append(Spacer(1, 0.2 * inch))
    
    # --- Datos de Alertas ---
    alertas = AlertaSeguridad.objects.select_related('residente_relacionado__user').all().order_by('-fecha_hora')[:30]
    
    data = [['Fecha', 'Tipo', 'Descripción', 'Residente', 'Estado']]
    
    for alerta in alertas:
        residente_str = alerta.residente_relacionado.user.get_full_name() if alerta.residente_relacionado else "N/A"
        estado_str = "RESUELTO" if alerta.resuelto else "PENDIENTE"
        
        # Usamos Paragraph para que el texto largo haga wrap en la celda
        desc_paragraph = Paragraph(alerta.descripcion, styles['Normal'])
        
        data.append([
            alerta.fecha_hora.strftime('%d/%m/%Y\n%H:%M'),
            alerta.get_tipo_alerta_display(),
            desc_paragraph,
            residente_str,
            estado_str
        ])

    # --- Estilos de Tabla ---
    table = Table(data, colWidths=[0.9*inch, 1.2*inch, 2.5*inch, 1.5*inch, 1.0*inch])
    
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (2, 1), (2, -1), 'LEFT'), # Descripción alineada a la izquierda
    ])
    
    # Colorear filas según estado (Iterar sobre data para aplicar estilo condicional es complejo en TableStyle estático, 
    # pero podemos colorear texto PENDIENTE en rojo)
    
    for i, row in enumerate(data[1:], 1):
        if row[-1] == "PENDIENTE":
            style.add('TEXTCOLOR', (4, i), (4, i), colors.red)
            style.add('FONTNAME', (4, i), (4, i), 'Helvetica-Bold')
        else:
            style.add('TEXTCOLOR', (4, i), (4, i), colors.green)
    
    table.setStyle(style)
    elements.append(table)
    
    # --- Pie de Página ---
    elements.append(Spacer(1, 0.5 * inch))
    elements.append(Paragraph("Generado automáticamente por SmartCondominium AI", styles['Italic']))

    doc.build(elements)
    buffer.seek(0)
    return buffer
